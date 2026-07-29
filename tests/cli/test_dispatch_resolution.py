"""The CLI resolves inputs from arguments, a named survey, or a prompt."""

from pathlib import Path

import openpyxl
import pytest

from gnb_survey.cli import actions
from gnb_survey.cli import dispatch as cli
from gnb_survey.triangulate.discovery import SurveyFiles

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"

_HEADER = (
    "Point Name,Code,Northing,Easting,Elevation,Latitude,Longitude,Altitude,"
    "Original Altitude,Measuring height,Antenna Height\n"
)
_ROWS = (
    "Pt1,,149955.9668,354634.9606,32.5729,1.35579855,103.69391447,32.5729,34.6389,2.0000,2.0660\n"
    "Pt2,,149962.1319,354630.4524,33.6284,1.35585427,103.69387394,33.6284,35.6944,2.0000,2.0660\n"
    "Pt3,,149965.0253,354619.5814,34.5987,1.35588038,103.69377627,34.5987,36.6647,2.0000,2.0660\n"
)


@pytest.fixture()
def data_root(tmp_path) -> Path:
    folder = tmp_path / "surveys" / "20260716" / "mappro"
    folder.mkdir(parents=True)
    (folder / "dd (Decimal).csv").write_text(_HEADER + _ROWS, encoding="latin-1")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Point Name", "distance", "angle", "height of binoc"])
    for row in (("pt1", 49.4, 51, 2.06), ("pt2", 54.0, 46, 2.08), ("pt3", 54.6, 46, 2.08)):
        sheet.append(list(row))
    workbook.save(tmp_path / "20260716_measurment_binoc.xlsx")
    return tmp_path


@pytest.fixture()
def data_root_without_binoc(tmp_path) -> Path:
    """A survey whose MapPro export exists but whose sightings haven't been typed up.

    Discovery now surfaces this survey rather than hiding it, so resolving it
    for solving must fail cleanly instead of crashing inside
    read_binoc_readings(None).
    """
    folder = tmp_path / "surveys" / "20260722" / "mappro"
    folder.mkdir(parents=True)
    (folder / "dd (Decimal).csv").write_text(_HEADER + _ROWS, encoding="latin-1")
    return tmp_path


@pytest.fixture()
def data_root_named_after_a_verb(tmp_path) -> Path:
    """A data root holding one ordinary survey and one folder called `solve`."""
    for name in ("20260716", "solve"):
        folder = tmp_path / "surveys" / name / "mappro"
        folder.mkdir(parents=True)
        (folder / "dd (Decimal).csv").write_text(_HEADER + _ROWS, encoding="latin-1")
    return tmp_path


def _run(argv, **kwargs):
    """Run the CLI with no terminal unless a test asks for one.

    `main` falls back to sys.stdin.isatty(), which is False under pytest's
    capture but True under `pytest -s`. Pinning it here keeps a suite run
    from blocking on the verb menu's `input()`.
    """
    kwargs.setdefault("is_tty", False)
    lines: list[str] = []
    code = cli.main(["survey.py", *argv], output_fn=lines.append, **kwargs)
    return code, "\n".join(lines)


@pytest.mark.unit
def test_list_prints_surveys_and_exits_zero(data_root):
    code, text = _run(["--list", "--data-root", str(data_root)])
    assert code == 0
    assert "20260716" in text


@pytest.mark.unit
def test_survey_by_name_solves_without_prompting(data_root):
    code, text = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_unknown_survey_lists_what_was_found(data_root):
    code, text = _run(["19990101", "--data-root", str(data_root)])
    assert code == 1
    assert "20260716" in text


# --- The Task 5 crash guard -------------------------------------------------
#
# A survey with MapPro exports but no <name>*.xlsx workbook must never reach
# read_binoc_readings(None). Review of Task 6 established that the guard now
# has THREE sites, not two, and that they are layered rather than parallel:
#
#   1. menu.select_verb refuses a blocked verb at the menu itself, so on the
#      prompted path site 2 below is never reached;
#   2. dispatch.main gates every verb through blocked_for -- this is what
#      catches the non-interactive path, and backs up site 1;
#   3. actions.do_solve refuses binoc=None on its own, because do_animate
#      calls it directly, passing through neither site above.
#
# One test each, below (site 1 is also covered by tests/cli/test_menu.py::
# test_verb_menu_shows_a_blocked_verb_with_its_reason). They look redundant
# and are not: deleting an upstream site is invisible in the exit code alone,
# because the next site down still returns 1. That is why each test asserts
# the *shape* of its own site's message rather than just `code == 1`.
#
# An earlier test that answered "" to both menus was deleted here: "" means
# `back` at the verb menu, so it exited at "cancelled." before any verb was
# attempted and passed even with all three sites removed.


@pytest.mark.unit
def test_named_survey_without_a_workbook_errors_instead_of_crashing(data_root_without_binoc):
    """Site 2. Discovery no longer excludes a binoc-less survey, so resolving
    it by name must fail cleanly rather than raising TypeError inside
    read_binoc_readings(None).

    "cannot solve" and "To fix:" are the capability gate's own shape. Without
    them the run still exits 1 -- do_solve's guard catches it -- so asserting
    the exit code alone would not distinguish site 2 from site 3.
    """
    code, text = _run(
        ["20260722", "--data-root", str(data_root_without_binoc)], is_tty=False
    )
    assert code == 1
    assert "cannot solve" in text
    assert "To fix:" in text
    assert "20260722" in text
    assert "xlsx" in text


@pytest.mark.unit
def test_prompted_solve_is_refused_when_the_workbook_is_missing(
    data_root_without_binoc,
):
    """Site 1, reached by picking the survey and then picking `solve`.

    The user answers "" (first survey), "2" (solve), then "b" to leave.
    "solve is unavailable" is printed only by select_verb's refusal, so this
    fails if the menu ever starts handing a blocked verb back to main.
    """
    answers = iter(["", "2", "b"])
    code, text = _run(
        ["--data-root", str(data_root_without_binoc)],
        is_tty=True,
        input_fn=lambda _: next(answers),
    )
    assert code == 1
    assert "solve is unavailable" in text
    assert "20260722" in text
    assert "xlsx" in text


@pytest.mark.unit
def test_do_solve_refuses_a_survey_with_no_workbook(tmp_path):
    """Site 3, reached by calling the action directly.

    `main` gates every verb through blocked_for, but do_solve is public and
    do_animate calls it, so it must refuse a binoc-less survey on its own
    rather than reaching read_binoc_readings(None).
    """
    mappro = tmp_path / "m.csv"
    files = SurveyFiles(
        name="20260722",
        mappro=mappro,
        exports=(mappro,),
        binoc=None,
        scene_json=None,
    )
    lines: list[str] = []

    code = actions.do_solve(
        files,
        cli._parse_args(["survey.py"]),
        output_dir=tmp_path / "out",
        output_fn=lines.append,
    )

    assert code == 1
    assert "no sightings workbook" in "\n".join(lines)
    assert not (tmp_path / "out").exists()


@pytest.mark.unit
def test_explicit_paths_are_named_after_the_parent_folder(data_root):
    survey = data_root / "surveys" / "20260716" / "mappro" / "dd (Decimal).csv"
    binoc = data_root / "20260716_measurment_binoc.xlsx"
    code, text = _run(["solve", str(survey), str(binoc)])
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_a_survey_name_and_explicit_paths_together_are_rejected(data_root):
    """`survey.py 20260716 solve A.csv B.xlsx` names its subject twice.

    Acting on the paths would drop the survey name, and acting on the survey
    would drop the paths -- either way the user typed something that was
    silently discarded. Both readings are named instead.
    """
    survey = data_root / "surveys" / "20260716" / "mappro" / "dd (Decimal).csv"
    binoc = data_root / "20260716_measurment_binoc.xlsx"
    code, text = _run(
        ["20260716", "solve", str(survey), str(binoc), "--data-root", str(data_root)]
    )
    assert code == 1
    assert "not both" in text
    assert "survey.py 20260716 solve" in text          # the survey reading
    assert f"survey.py solve {survey}" in text         # the explicit-paths reading


@pytest.mark.unit
def test_name_option_overrides_the_survey_name(data_root):
    code, text = _run(["20260716", "--data-root", str(data_root), "--name", "Cetran"])
    assert code == 0
    assert "Cetran gNB" in text


@pytest.mark.unit
def test_missing_explicit_path_errors_rather_than_prompting(data_root):
    code, text = _run(["solve", "/no/such/survey.csv", "/no/such/binoc.xlsx"])
    assert code == 1
    assert "not a file" in text


@pytest.mark.unit
def test_non_interactive_refuses_to_prompt(data_root):
    code, text = _run(["--data-root", str(data_root), "--non-interactive"], is_tty=True)
    assert code == 1
    assert "survey name" in text


@pytest.mark.unit
def test_no_input_is_the_same_flag_under_its_new_name(data_root):
    """--non-interactive survives as an alias; --no-input is the primary spelling."""
    code, text = _run(["--data-root", str(data_root), "--no-input"], is_tty=True)
    assert code == 1
    assert "survey name" in text


@pytest.mark.unit
def test_not_a_terminal_refuses_to_prompt(data_root):
    code, text = _run(["--data-root", str(data_root)], is_tty=False)
    assert code == 1
    assert "survey name" in text


@pytest.mark.unit
def test_prompt_is_used_when_attached_to_a_terminal(data_root):
    """Enter takes the first survey, then 2 picks solve from the verb menu."""
    answers = iter(["", "2"])
    code, text = _run(
        ["--data-root", str(data_root)],
        is_tty=True,
        input_fn=lambda _: next(answers),
    )
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_aborting_the_prompt_exits_without_a_traceback(data_root):
    def refuse(_prompt):
        raise EOFError

    code, _ = _run(["--data-root", str(data_root)], is_tty=True, input_fn=refuse)
    assert code == 1


@pytest.mark.unit
def test_no_surveys_found_names_the_expected_layout(tmp_path):
    code, text = _run(["--data-root", str(tmp_path), "--list"])
    assert code == 1
    assert "surveys" in text


@pytest.mark.unit
def test_bad_field_data_ends_the_run_even_when_interactive(data_root):
    """Re-prompting cannot repair a spreadsheet, so this exits rather than looping."""
    workbook = openpyxl.load_workbook(data_root / "20260716_measurment_binoc.xlsx")
    sheet = workbook.active
    sheet.cell(row=4, column=1, value="pt1")  # duplicate of row 2
    workbook.save(data_root / "20260716_measurment_binoc.xlsx")

    answers = iter(["", "2"])
    code, text = _run(
        ["--data-root", str(data_root)],
        is_tty=True,
        input_fn=lambda _: next(answers),
    )
    assert code == 1
    assert "twice" in text


@pytest.mark.unit
def test_sigma_overrides_reach_the_report(data_root):
    code, text = _run(
        ["20260716", "--data-root", str(data_root),
         "--sigma-distance", "1.0", "--sigma-elevation", "0.5"]
    )
    assert code == 0
    assert "1.00 m / 0.50°" in text


@pytest.mark.integration
def test_real_survey_end_to_end():
    if not FIXTURES.is_dir():
        pytest.skip("fixtures data not present")
    code, text = _run(["20260716", "--data-root", str(FIXTURES)])
    assert code == 0
    assert "1.35554" in text


@pytest.mark.unit
def test_csv_flag_writes_the_file(data_root, tmp_path):
    out = tmp_path / "gnb.csv"
    code, text = _run(
        ["20260716", "--data-root", str(data_root), "--csv", str(out)]
    )
    assert code == 0
    assert out.is_file()
    assert str(out) in text


@pytest.mark.unit
def test_csv_has_the_gnb_and_a_full_ring(data_root, tmp_path):
    import csv as _csv

    from gnb_survey.triangulate.mymaps import RING_POINTS

    out = tmp_path / "gnb.csv"
    _run(["20260716", "--data-root", str(data_root), "--csv", str(out)])
    with out.open(newline="", encoding="utf-8") as handle:
        rows = list(_csv.DictReader(handle))
    assert len(rows) == RING_POINTS + 1
    assert rows[0]["Code"] == "gNB"
    assert rows[1]["Code"] == "ellipse95"


@pytest.mark.unit
def test_csv_to_a_missing_directory_names_the_directory(data_root, tmp_path):
    out = tmp_path / "nope" / "gnb.csv"
    code, text = _run(
        ["20260716", "--data-root", str(data_root), "--csv", str(out)]
    )
    assert code == 1
    assert "directory" in text


@pytest.mark.unit
def test_csv_is_written_by_default_into_output(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    code, text = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    written = tmp_path / "output" / "20260716_gnb.csv"
    assert written.is_file()
    assert str(written) in text


@pytest.mark.unit
def test_default_output_directory_is_created_when_absent(data_root, tmp_path, monkeypatch):
    target = tmp_path / "deep" / "output"
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", target)
    code, _ = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    assert (target / "20260716_gnb.csv").is_file()


@pytest.mark.unit
def test_no_csv_suppresses_the_default(data_root, tmp_path, monkeypatch):
    """The folder itself still appears -- it is where the scene data goes."""
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    code, _ = _run(["20260716", "--data-root", str(data_root), "--no-csv"])
    assert code == 0
    assert not (tmp_path / "output" / "20260716_gnb.csv").exists()


@pytest.mark.unit
def test_explicit_csv_still_wins_over_the_default(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    chosen = tmp_path / "mine.csv"
    code, _ = _run(["20260716", "--data-root", str(data_root), "--csv", str(chosen)])
    assert code == 0
    assert chosen.is_file()
    assert not (tmp_path / "output" / "20260716_gnb.csv").exists()


@pytest.mark.unit
def test_explicit_csv_into_a_missing_directory_still_errors(data_root, tmp_path, monkeypatch):
    """An explicit path is a typo risk; only the default directory self-creates."""
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    code, text = _run(
        ["20260716", "--data-root", str(data_root), "--csv", str(tmp_path / "nope" / "x.csv")]
    )
    assert code == 1
    assert "directory" in text


@pytest.mark.unit
def test_bare_name_selects_a_survey(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "out")
    code, text = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_a_lone_file_path_says_to_put_the_verb_first(data_root, tmp_path, monkeypatch):
    """A single argument that is a file is a half-typed command, not a survey."""
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "out")
    survey = data_root / "surveys" / "20260716" / "mappro" / "dd (Decimal).csv"
    code, text = _run([str(survey), "--data-root", str(data_root)])
    assert code == 1
    assert "verb first" in text


@pytest.mark.unit
def test_a_survey_folder_named_after_a_verb_is_rejected(data_root_named_after_a_verb):
    """`survey.py solve` would mean both "solve something" and "the survey `solve`".

    The check is global, not per-command: one such folder blocks commands for
    every survey, including the ones whose names are unambiguous. That is
    deliberate -- the fix is to rename the folder, and an error that appeared
    only when the colliding name was typed would be easy to route around and
    leave in place for months.
    """
    code, text = _run(["20260716", "--data-root", str(data_root_named_after_a_verb)])
    assert code == 1
    assert "named after verbs" in text
    assert "solve" in text


@pytest.mark.unit
def test_list_survives_a_survey_folder_named_after_a_verb(data_root_named_after_a_verb):
    """--list returns before the collision check, so the diagnosis path still works.

    The collision error names the offending folder, but the user still needs
    to see the whole data root to fix it -- and --list is the only command
    left that runs.
    """
    code, text = _run(["--list", "--data-root", str(data_root_named_after_a_verb)])
    assert code == 0
    assert "solve" in text
    assert "20260716" in text


@pytest.mark.unit
def test_survey_flag_is_gone(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "out")
    with pytest.raises(SystemExit):
        _run(["--survey", "20260716", "--data-root", str(data_root)])
