"""The CLI resolves inputs from arguments, a named campaign, or a prompt."""

from pathlib import Path

import openpyxl
import pytest

import main as cli

RAW = Path(__file__).resolve().parents[2] / "raw_data"

_HEADER = (
    "Point Name,Code,Northing,Easting,Elevation,Latitude,Longitude,Altitude,"
    "Original Altitude,Measuring height,Antenna Height\n"
)
_ROWS = (
    "Pt1,,149955.9668,354634.9606,32.5729,1.35579855,103.69391447,32.5729,34.6389,2.0000,2.0660\n"
    "Pt2,,149962.1319,354630.4524,33.6284,1.35585427,103.69387394,33.6284,35.6944,2.0000,2.0660\n"
    "Pt3,,149965.0253,354619.5814,34.5987,1.35588038,103.69377627,34.5987,36.6647,2.0000,2.0660\n"
)


@pytest.fixture()
def data_root(tmp_path) -> Path:
    folder = tmp_path / "map-pro-csv" / "20260716"
    folder.mkdir(parents=True)
    (folder / "dd (Decimal).csv").write_text(_HEADER + _ROWS, encoding="latin-1")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Point Name", "distance", "angle", "height of binoc"])
    for row in (("pt1", 49.4, 51, 2.06), ("pt2", 54.0, 46, 2.08), ("pt3", 54.6, 46, 2.08)):
        sheet.append(list(row))
    workbook.save(tmp_path / "20260716_measurment_binoc.xlsx")
    return tmp_path


def _run(argv, **kwargs):
    lines: list[str] = []
    code = cli.main(["main.py", *argv], output_fn=lines.append, **kwargs)
    return code, "\n".join(lines)


@pytest.mark.unit
def test_list_prints_campaigns_and_exits_zero(data_root):
    code, text = _run(["--list", "--data-root", str(data_root)])
    assert code == 0
    assert "20260716" in text


@pytest.mark.unit
def test_campaign_by_name_solves_without_prompting(data_root):
    code, text = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_unknown_campaign_lists_what_was_found(data_root):
    code, text = _run(["19990101", "--data-root", str(data_root)])
    assert code == 1
    assert "20260716" in text


@pytest.mark.unit
def test_explicit_paths_are_named_after_the_parent_folder(data_root):
    survey = data_root / "map-pro-csv" / "20260716" / "dd (Decimal).csv"
    binoc = data_root / "20260716_measurment_binoc.xlsx"
    code, text = _run([str(survey), str(binoc)])
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_name_option_overrides_the_campaign_name(data_root):
    code, text = _run(["20260716", "--data-root", str(data_root), "--name", "Cetran"])
    assert code == 0
    assert "Cetran gNB" in text


@pytest.mark.unit
def test_missing_explicit_path_errors_rather_than_prompting(data_root):
    code, text = _run(["/no/such/survey.csv", "/no/such/binoc.xlsx"])
    assert code == 1
    assert "not found" in text


@pytest.mark.unit
def test_non_interactive_refuses_to_prompt(data_root):
    code, text = _run(["--data-root", str(data_root), "--non-interactive"])
    assert code == 1
    assert "campaign name" in text


@pytest.mark.unit
def test_not_a_terminal_refuses_to_prompt(data_root):
    code, text = _run(["--data-root", str(data_root)], is_tty=False)
    assert code == 1
    assert "campaign name" in text


@pytest.mark.unit
def test_prompt_is_used_when_attached_to_a_terminal(data_root):
    answers = iter([""])
    code, text = _run(
        ["--data-root", str(data_root)],
        is_tty=True,
        input_fn=lambda _: next(answers),
    )
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_aborting_the_prompt_exits_without_a_traceback(data_root):
    def refuse(_prompt):
        raise EOFError

    code, _ = _run(["--data-root", str(data_root)], is_tty=True, input_fn=refuse)
    assert code == 1


@pytest.mark.unit
def test_no_campaigns_found_names_the_expected_layout(tmp_path):
    code, text = _run(["--data-root", str(tmp_path), "--list"])
    assert code == 1
    assert "map-pro-csv" in text


@pytest.mark.unit
def test_bad_field_data_ends_the_run_even_when_interactive(data_root):
    """Re-prompting cannot repair a spreadsheet, so this exits rather than looping."""
    workbook = openpyxl.load_workbook(data_root / "20260716_measurment_binoc.xlsx")
    sheet = workbook.active
    sheet.cell(row=4, column=1, value="pt1")  # duplicate of row 2
    workbook.save(data_root / "20260716_measurment_binoc.xlsx")

    code, text = _run(
        ["--data-root", str(data_root)], is_tty=True, input_fn=lambda _: ""
    )
    assert code == 1
    assert "twice" in text


@pytest.mark.unit
def test_sigma_overrides_reach_the_report(data_root):
    code, text = _run(
        ["20260716", "--data-root", str(data_root),
         "--sigma-distance", "1.0", "--sigma-elevation", "0.5"]
    )
    assert code == 0
    assert "1.00 m / 0.50°" in text


@pytest.mark.integration
def test_real_campaign_end_to_end():
    if not RAW.is_dir():
        pytest.skip("raw data not present")
    code, text = _run(["20260716", "--data-root", str(RAW)])
    assert code == 0
    assert "1.35554" in text


@pytest.mark.unit
def test_csv_flag_writes_the_file(data_root, tmp_path):
    out = tmp_path / "gnb.csv"
    code, text = _run(
        ["20260716", "--data-root", str(data_root), "--csv", str(out)]
    )
    assert code == 0
    assert out.is_file()
    assert str(out) in text


@pytest.mark.unit
def test_csv_has_the_gnb_and_a_full_ring(data_root, tmp_path):
    import csv as _csv

    from gnb_triangulate.mymaps import RING_POINTS

    out = tmp_path / "gnb.csv"
    _run(["20260716", "--data-root", str(data_root), "--csv", str(out)])
    with out.open(newline="", encoding="utf-8") as handle:
        rows = list(_csv.DictReader(handle))
    assert len(rows) == RING_POINTS + 1
    assert rows[0]["Code"] == "gNB"
    assert rows[1]["Code"] == "ellipse95"


@pytest.mark.unit
def test_csv_to_a_missing_directory_names_the_directory(data_root, tmp_path):
    out = tmp_path / "nope" / "gnb.csv"
    code, text = _run(
        ["20260716", "--data-root", str(data_root), "--csv", str(out)]
    )
    assert code == 1
    assert "directory" in text


@pytest.mark.unit
def test_csv_is_written_by_default_into_output(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    code, text = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    written = tmp_path / "output" / "20260716_gnb.csv"
    assert written.is_file()
    assert str(written) in text


@pytest.mark.unit
def test_default_output_directory_is_created_when_absent(data_root, tmp_path, monkeypatch):
    target = tmp_path / "deep" / "output"
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", target)
    code, _ = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    assert (target / "20260716_gnb.csv").is_file()


@pytest.mark.unit
def test_no_csv_suppresses_the_default(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    code, _ = _run(["20260716", "--data-root", str(data_root), "--no-csv"])
    assert code == 0
    assert not (tmp_path / "output").exists()


@pytest.mark.unit
def test_explicit_csv_still_wins_over_the_default(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    chosen = tmp_path / "mine.csv"
    code, _ = _run(["20260716", "--data-root", str(data_root), "--csv", str(chosen)])
    assert code == 0
    assert chosen.is_file()
    assert not (tmp_path / "output").exists()


@pytest.mark.unit
def test_explicit_csv_into_a_missing_directory_still_errors(data_root, tmp_path, monkeypatch):
    """An explicit path is a typo risk; only the default directory self-creates."""
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "output")
    code, text = _run(
        ["20260716", "--data-root", str(data_root), "--csv", str(tmp_path / "nope" / "x.csv")]
    )
    assert code == 1
    assert "directory" in text


@pytest.mark.unit
def test_bare_name_selects_a_campaign(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "out")
    code, text = _run(["20260716", "--data-root", str(data_root)])
    assert code == 0
    assert "20260716 gNB" in text


@pytest.mark.unit
def test_a_lone_file_path_asks_for_the_second_file(data_root, tmp_path, monkeypatch):
    """A single argument that is a file is a half-typed command, not a campaign."""
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "out")
    survey = data_root / "map-pro-csv" / "20260716" / "dd (Decimal).csv"
    code, text = _run([str(survey), "--data-root", str(data_root)])
    assert code == 1
    assert "both" in text


@pytest.mark.unit
def test_campaign_flag_is_gone(data_root, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_DEFAULT_OUTPUT_DIR", tmp_path / "out")
    with pytest.raises(SystemExit):
        _run(["--campaign", "20260716", "--data-root", str(data_root)])
